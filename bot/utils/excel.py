import io
import calendar
from datetime import date
from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_

from bot.database.models import Report, User, Plan, ManagementExpense, Project


# --- Palette -----------------------------------------------------------------

_F_BLACK  = Font(bold=False, color="000000", size=9)
_F_BOLD   = Font(bold=True,  color="000000", size=9)
_F_WHITE  = Font(bold=True,  color="FFFFFF", size=9)
_F_GREEN  = Font(bold=True,  color="339933", size=9)
_F_RED    = Font(bold=True,  color="FF0000", size=9)
_F_BLUE   = Font(bold=True,  color="0000FF", size=9)

_FILL_PROJECT   = PatternFill("solid", fgColor="B8CCE4")   # light blue
_FILL_GREEN     = PatternFill("solid", fgColor="C4D79B")   # light green 
_FILL_BLUE      = PatternFill("solid", fgColor="95B3D7")   # blue headers
_FILL_GRAY      = PatternFill("solid", fgColor="F2F2F2")   # alt row
_FILL_RED_HDR   = PatternFill("solid", fgColor="FF0000")   # Red header for ИТОГО
_FILL_BLUE_IN   = PatternFill("solid", fgColor="9DC3E6")   # total row color
_FILL_WHITE     = PatternFill("solid", fgColor="FFFFFF")   # white background for days

_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Border styles
_thin  = Side(border_style="thin",   color="000000")
_thick = Side(border_style="thick",  color="000000")
_med   = Side(border_style="medium", color="000000")

_BORDER       = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thin)
_BORDER_THICK = Border(left=_thick, right=_thick, top=_thick, bottom=_thick)
_BORDER_HDR   = Border(left=_thick, right=_thick, top=_thick, bottom=_med)
_BORDER_BOT   = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_med)
_BORDER_LEFT_THICK  = Border(left=_thick, right=_thin, top=_thin, bottom=_thin)
_BORDER_RIGHT_THICK = Border(left=_thin,  right=_thick, top=_thin, bottom=_thin)
_BORDER_TOP_MED     = Border(left=_thin,  right=_thin,  top=_med,  bottom=_thin)
_BORDER_BOT_MED     = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_med)

_NUM_FMT = '#,##0.00 "₽"'
_INT_FMT = '#,##0 "₽"'
_PCT_FMT = '0%'


def _apply_border(ws, r1, c1, r2, c2, border_style="medium"):
    """Apply a consistent border around a rectangular region of cells."""
    side = {"thin": _thin, "medium": _med, "thick": _thick}.get(border_style, _med)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            left   = side if c == c1 else cell.border.left
            right  = side if c == c2 else cell.border.right
            top    = side if r == r1 else cell.border.top
            bottom = side if r == r2 else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _cell(ws, row, col, value="", fill=None, font=None, align=None, fmt=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill = fill
    if font:   c.font = font or _F_BLACK
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    c.border = border or _BORDER
    return c


def _merge(ws, r1, c1, r2, c2, value="", fill=None, font=None, align=None, fmt=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if fill:  c.fill = fill
    if font:  c.font = font
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    _b = border or _BORDER
    for r in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            ws.cell(row=r, column=col).border = _b
    return c


def _month_label(m: int) -> str:
    return ["ЯНВ", "ФЕВ", "МАР", "АПР", "МАЙ", "ИЮН", "ИЮЛ", "АВГ", "СЕН", "ОКТ", "НОЯ", "ДЕК"][m-1]


async def generate_monthly_calendar(
    session: AsyncSession,
    year: int,
    month: int,
    city: str = "all"
) -> bytes:
    _, days_in_month = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end   = date(year, month, days_in_month)

    # --- Fetch all relevant data once ----------------------------------------

    q = select(Report).where(Report.date >= start, Report.date <= end)
    if city != "all": q = q.where(or_(Report.city == city, Report.city == None))
    res = await session.execute(q.order_by(Report.project_name, Report.date))
    all_reports = res.scalars().all()

    q_plans = select(Plan).where(Plan.is_active == True, Plan.period == "month")
    if city != "all": q_plans = q_plans.where(Plan.city == city)
    all_plans = (await session.execute(q_plans)).scalars().all()

    mq = select(ManagementExpense).where(ManagementExpense.date >= start, ManagementExpense.date <= end)
    if city != "all": mq = mq.where(ManagementExpense.city == city)
    all_mgmt = (await session.execute(mq)).scalars().all()

    q_proj = select(Project).where(Project.is_active == True)
    if city != "all": q_proj = q_proj.where(Project.city == city)
    all_projects = (await session.execute(q_proj)).scalars().all()

    cities_to_process = ["gomel", "minsk"] if city == "all" else [city]

    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "Доходы", "нал.", "безнал.", "Расходы",
        "зарплата\nфотографа", "зарплата\nстажера", "хоз расход", 
        "расходник", "УСН 6%", "налоги по\nЗП 35,6%", "техника", "аренда", 
        "Остаток конец дня", "из них нал."
    ]

    def build_city_sheet(sheet_city: str, reports: list[Report], plans: list[Plan], mgmt_list: list[ManagementExpense]):
        city_label = {"gomel": "Гомель", "minsk": "Минск"}.get(sheet_city, sheet_city.title())
        ws = wb.create_sheet(title=city_label)
        
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 18
        for i in range(14):
            ws.column_dimensions[get_column_letter(5 + i)].width = 15

        row = 1
        plan_by_project = {p.project_name: p.plan_amount for p in plans if p.project_name}
        
        # Valid active projects
        active_proj_names = {p.name for p in all_projects if p.city == sheet_city}
        project_set = set()
        # Include projects that have actual reports this month
        project_set.update(r.project_name for r in reports if r.city == sheet_city or r.city is None)
        # Include projects that both have an active plan AND are currently an active project
        project_set.update(p.project_name for p in plans if p.project_name and p.city == sheet_city and p.project_name in active_proj_names)
        projects_sorted = sorted(project_set)

        by_project = defaultdict(lambda: defaultdict(list))
        for r in reports:
            if r.city == sheet_city or r.city is None:
                by_project[r.project_name][r.date.day].append(r)

        # Totals for City Level
        c_plan = c_rev = c_cash = c_acq = c_exp = c_grand_exp = 0.0
        c_sal = c_tra = 0.0
        c_cons = c_usn = c_tax = c_tech = c_rent = 0.0

        for p_name in projects_sorted:
            p_data = by_project[p_name]
            p_plan = plan_by_project.get(p_name, 0)
            
            linked_mgmt = [m for m in mgmt_list if m.project_name == p_name]
            
            agg = {}
            total_rev = total_cash = total_acq = total_exp = 0.0
            total_sal = total_tra = 0.0
            total_auto_usn = total_auto_tax_zp = 0.0
            
            p_cons = p_usn = p_tax = p_tech = p_rent = 0.0
            
            for d in range(1, days_in_month + 1):
                reps = p_data.get(d, [])
                master = max(reps, key=lambda r: float(r.revenue)) if reps else None
                day_rev = float(master.revenue) if master else 0.0
                day_cash = float(master.cash) if master else 0.0
                day_acq = float(master.acquiring) if master else 0.0
                day_exp = float(master.expense) if master else 0.0
                
                day_sal_total = sum(float(r.salary_paid) for r in reps)
                day_tra_total = sum(float(r.trainee_salary) for r in reps)
                
                day_auto_usn = day_rev * 0.06
                day_auto_tax_zp = (day_sal_total + day_tra_total) * 0.356
                
                # Fetch mgmt expenses recorded SPECIFICALLY on this day!
                def _d_sum(cat): return float(sum(m.amount for m in linked_mgmt if m.category == cat and m.date.day == d))
                
                day_cons = _d_sum("расходник")
                day_usn  = day_auto_usn + _d_sum("усн_6")
                day_tax  = day_auto_tax_zp + _d_sum("налоги_зп")
                day_tech = _d_sum("техника")
                day_rent = _d_sum("аренда")
                
                day_all_mgmt = day_cons + day_usn + day_tax + day_tech + day_rent
                day_total_dist_exp = day_sal_total + day_tra_total + day_exp + day_all_mgmt
                
                agg[d] = {
                    "reps": reps,
                    "rev": day_rev,
                    "cash": day_cash,
                    "acq": day_acq,
                    "exp": day_exp,
                    "sal_total": day_sal_total,
                    "tra_total": day_tra_total,
                    "cons": day_cons,
                    "usn": day_usn,
                    "tax": day_tax,
                    "tech": day_tech,
                    "rent": day_rent,
                    "total_exp": day_total_dist_exp,
                    "ostatok": day_rev - day_total_dist_exp,
                    "incass": day_cash - day_exp - day_tra_total,
                }
                
                total_rev += day_rev
                total_cash += day_cash
                total_acq += day_acq
                total_exp += day_exp
                total_sal += day_sal_total
                total_tra += day_tra_total
                total_auto_usn += day_auto_usn
                total_auto_tax_zp += day_auto_tax_zp
                
                p_cons += day_cons
                p_usn += day_usn
                p_tax += day_tax
                p_tech += day_tech
                p_rent += day_rent

            total_pct = (total_rev / p_plan) if p_plan > 0 else 0.0

            proj_start_row = row  # Track start for outer thick border

            # Row 1: Top header with "в нал" labels
            ws.row_dimensions[row].height = 18
            _merge(ws, row, 1, row, 2, p_name, fill=_FILL_PROJECT, font=_F_BOLD, align=_CTR)
            _cell(ws, row, 3, "", fill=_FILL_PROJECT)
            _cell(ws, row, 4, "", fill=_FILL_PROJECT)
            for i in range(14):
                if headers[i] in ["зарплата\nфотографа", "зарплата\nстажера"]:
                    _cell(ws, row, 5 + i, "в нал", fill=_FILL_PROJECT, align=_CTR, font=_F_RED)
                else:
                    _cell(ws, row, 5 + i, "", fill=_FILL_PROJECT)
            row += 1

            # Row 2: Headers
            ws.row_dimensions[row].height = 32
            _cell(ws, row, 1, "План", fill=_FILL_PROJECT, align=_LEFT)
            _cell(ws, row, 2, float(p_plan), fill=_FILL_PROJECT, fmt=_INT_FMT, align=_CTR)
            _cell(ws, row, 3, "ДАТА", fill=_FILL_BLUE, font=_F_BOLD, align=_CTR)
            _cell(ws, row, 4, "ФИО", fill=_FILL_BLUE, font=_F_BOLD, align=_CTR)
            for i, h in enumerate(headers):
                c = _cell(ws, row, 5 + i, h, fill=_FILL_BLUE, align=_CTR)
                if h in ["Доходы", "зарплата\nфотографа", "Остаток конец дня", "из них нал."]:
                    c.font = _F_GREEN
                elif h in ["нал.", "безнал.", "зарплата\nстажера"]:
                    c.font = _F_RED
                else:
                    c.font = _F_BLUE
            # Medium bottom border under the header row
            _apply_border(ws, row, 1, row, 18, "medium")
            row += 1

            # Row 3: Total ("Общая")
            grand_total_exp = total_sal + total_tra + total_exp + (p_cons + p_usn + p_tax + p_tech + p_rent)
            
            # Append city totals
            c_plan += p_plan
            c_rev += total_rev
            c_cash += total_cash
            c_acq += total_acq
            c_exp += total_exp
            c_sal += total_sal
            c_tra += total_tra
            c_cons += p_cons
            c_usn += p_usn
            c_tax += p_tax
            c_tech += p_tech
            c_rent += p_rent
            c_grand_exp += grand_total_exp

            ws.row_dimensions[row].height = 18
            _cell(ws, row, 1, "Выполнено", fill=_FILL_PROJECT, align=_LEFT)
            _cell(ws, row, 2, total_pct, fill=_FILL_PROJECT, fmt=_PCT_FMT, align=_CTR)
            _cell(ws, row, 3, "", fill=_FILL_GRAY)
            _cell(ws, row, 4, "Общая", fill=_FILL_GRAY, font=_F_BOLD, align=_CTR)
            _cell(ws, row, 5, total_rev, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 6, total_cash, fill=_FILL_BLUE_IN, fmt=_NUM_FMT) 
            _cell(ws, row, 7, total_acq, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 8, grand_total_exp, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 9, total_sal, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 10, total_tra, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 11, total_exp, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 12, p_cons, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 13, p_usn, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 14, p_tax, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 15, p_tech, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 16, p_rent, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 17, total_rev - grand_total_exp, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            _cell(ws, row, 18, total_cash - total_exp - total_tra, fill=_FILL_BLUE_IN, fmt=_NUM_FMT)
            # Medium bottom of totals row
            _apply_border(ws, row, 1, row, 18, "medium")
            row += 1

            days_start_row = row  # Start of daily rows

            # Row 4 to N: Days
            for d in range(1, days_in_month + 1):
                day_data = agg[d]
                reps = day_data["reps"]
                n_rows = max(1, len(reps))
                start_r = row
                end_r = row + n_rows - 1

                # Alternate row background for readability
                row_fill = _FILL_WHITE if d % 2 == 0 else _FILL_GRAY

                _merge(ws, start_r, 1, end_r, 1, "", fill=row_fill, align=_CTR)
                _merge(ws, start_r, 2, end_r, 2, "", fill=row_fill, align=_CTR)

                date_label = f"{d} {_month_label(month)}"
                _merge(ws, start_r, 3, end_r, 3, date_label, fill=row_fill, align=_CTR, font=_F_BOLD)
                
                def _num(val): return val if val != 0 else ""

                _merge(ws, start_r, 5, end_r, 5, _num(day_data["rev"]), fill=_FILL_GREEN, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 6, end_r, 6, _num(day_data["cash"]), fill=row_fill, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 7, end_r, 7, _num(day_data["acq"]), fill=row_fill, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 8, end_r, 8, _num(day_data["total_exp"]), fill=_FILL_GREEN, fmt=_NUM_FMT, align=_CTR)
                
                if not reps:
                    _cell(ws, start_r, 4, "", fill=row_fill)
                    _cell(ws, start_r, 9, "", fill=row_fill, fmt=_NUM_FMT)
                    _cell(ws, start_r, 10, "", fill=row_fill, fmt=_NUM_FMT)
                else:
                    for i, r in enumerate(reps):
                        cur_r = start_r + i
                        fname = r.employee_name.split()[0] if r.employee_name else "Unknown"
                        _cell(ws, cur_r, 4, fname, fill=row_fill, align=_CTR)
                        _cell(ws, cur_r, 9, _num(float(r.salary_paid)), fill=row_fill, fmt=_NUM_FMT)
                        _cell(ws, cur_r, 10, _num(float(r.trainee_salary)), fill=row_fill, fmt=_NUM_FMT)
                
                _merge(ws, start_r, 11, end_r, 11, _num(day_data["exp"]), fill=row_fill, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 12, end_r, 12, _num(day_data["cons"]), fill=row_fill, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 13, end_r, 13, _num(day_data["usn"]), fill=row_fill, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 14, end_r, 14, _num(day_data["tax"]), fill=row_fill, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 15, end_r, 15, _num(day_data["tech"]), fill=row_fill, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 16, end_r, 16, _num(day_data["rent"]), fill=row_fill, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 17, end_r, 17, _num(day_data["ostatok"]), fill=_FILL_GREEN, fmt=_NUM_FMT, align=_CTR)
                _merge(ws, start_r, 18, end_r, 18, _num(day_data["incass"]), fill=_FILL_GREEN, fmt=_NUM_FMT, align=_CTR)
                
                row = end_r + 1
            
            # Apply thick outer border around the FULL project block (title + header + total + days)
            _apply_border(ws, proj_start_row, 1, row - 1, 18, "thick")
            # Medium separators at key horizontal boundaries within the block
            _apply_border(ws, proj_start_row, 1, proj_start_row, 18, "medium")  # title bottom
            _apply_border(ws, days_start_row - 1, 1, days_start_row - 1, 18, "medium")  # below total
            
            row += 2  # gap between projects

        if len(projects_sorted) > 0:
            row += 1  # Add a tiny gap before total
            _merge(ws, row, 1, row, 18, f"ИТОГО ПО ВСЕМ ПРОЕКТАМ — {city_label}", fill=_FILL_RED_HDR, font=_F_WHITE, align=_CTR)
            row += 1
            
            start_row = row
            
            # Row A-B: План and Выполнено
            _cell(ws, start_row, 1, "План", font=_F_BOLD)
            _cell(ws, start_row, 2, float(c_plan), fmt=_INT_FMT, align=_CTR)
            
            c_pct = (c_rev / c_plan) if c_plan > 0 else 0.0
            _cell(ws, start_row + 1, 1, "Выполнено", font=_F_RED)
            _cell(ws, start_row + 1, 2, c_pct, fill=_FILL_GREEN, fmt=_PCT_FMT, align=_CTR)

            # Row C-D: Vertical metrics list
            metrics = [
                ("Доходы", c_rev, _FILL_GREEN),
                ("нал.", c_cash, None),
                ("безнал.", c_acq, None),
                ("Расходы", c_grand_exp, _FILL_GREEN),
                ("зарплата Фотографа", c_sal, None),
                ("зарплата Стажера", c_tra, None),
                ("хоз расход", c_exp, None),
                ("расходник", c_cons, None),
                ("УСН 6%", c_usn, None),
                ("налоги по ЗП 35,6%", c_tax, None),
                ("техника", c_tech, None),
                ("аренда", c_rent, None),
                ("другое", 0.0, None), 
                ("Остаток конец дня", c_rev - c_grand_exp, _FILL_GREEN),
                ("из них нал.", c_cash - c_exp - c_tra, _FILL_GREEN),
            ]
            
            for i, (name, val, fill) in enumerate(metrics):
                _cell(ws, start_row + i, 3, name, fill=fill, font=_F_BOLD if fill else _F_BLACK)
                _cell(ws, start_row + i, 4, float(val), fill=fill, fmt=_NUM_FMT)
            
            totals_end_row = start_row + len(metrics) - 1
            # Thick outer border around the full city totals block
            _apply_border(ws, start_row - 1, 1, totals_end_row, 4, "thick")
            # Freeze panes: keep first 3 rows and 4 columns visible on scroll
            ws.freeze_panes = "E4"
            
            row = totals_end_row + 2

    for c_id in cities_to_process:
        city_reports = [r for r in all_reports if r.city == c_id or r.city is None]
        city_plans   = [p for p in all_plans if p.city == c_id]
        city_mgmt    = [m for m in all_mgmt if m.city == c_id]
        # if there are any reports or plans, build the sheet
        if city_reports or city_plans:
            build_city_sheet(c_id, city_reports, city_plans, city_mgmt)

    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="Нет Данных")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def generate_excel_report(session: AsyncSession, start_date: date, end_date: date) -> bytes:
    from openpyxl import Workbook as WB
    res = await session.execute(
        select(Report, User)
        .join(User, Report.user_id == User.id)
        .where(Report.date >= start_date, Report.date <= end_date)
        .order_by(Report.date, Report.project_name)
    )
    rows = res.all()

    wb2 = WB()
    ws2 = wb2.active
    ws2.title = f"Отчет {start_date.strftime('%d.%m')}-{end_date.strftime('%d.%m.%Y')}"

    H_FONT = Font(bold=True, color="FFFFFF", size=11)
    H_FILL = PatternFill("solid", fgColor="1F4E79")
    ALT    = PatternFill("solid", fgColor="DEEAF1")
    TOT_F  = Font(bold=True, color="FFFFFF", size=11)
    TOT_FL = PatternFill("solid", fgColor="2E75B6")
    CENTER = Alignment(horizontal="center", vertical="center")
    thin   = Side(border_style="thin", color="AAAAAA")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        ("Дата", 13), ("Проект", 22), ("Сотрудник", 20), ("Чел.", 7),
        ("Выручка", 13), ("Нал", 13), ("Безнал", 13), ("ЗП", 13),
        ("Расход", 13), ("Остаток", 14), ("Посетит.", 9), ("ДР", 6), ("Комментарий", 30),
    ]
    for col, (h, w) in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = H_FONT, H_FILL, CENTER, BORDER
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 24

    totals = dict(revenue=0.0, cash=0.0, acq=0.0, salary=0.0, expense=0.0, visitors=0, bdays=0)
    for ri, (rep, user) in enumerate(rows, 2):
        fill = ALT if ri % 2 == 0 else None
        data = [
            rep.date.strftime("%d.%m.%Y"), rep.project_name, rep.employee_name,
            rep.shift_count, rep.revenue, rep.cash, rep.acquiring,
            rep.salary_paid, rep.expense, rep.cash_balance,
            rep.visitors, rep.birthdays, rep.comment or "",
        ]
        for ci, v in enumerate(data, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            if fill: c.fill = fill
            c.border = BORDER
            if ci in (1, 4, 11, 12): c.alignment = CENTER
        totals["revenue"]  += rep.revenue
        totals["cash"]     += rep.cash
        totals["acq"]      += rep.acquiring
        totals["salary"]   += rep.salary_paid
        totals["expense"]  += rep.expense
        totals["visitors"] += rep.visitors
        totals["bdays"]    += rep.birthdays

    tr = len(rows) + 2
    summary = ["ИТОГО", "", "", "", totals["revenue"], totals["cash"], totals["acq"],
                totals["salary"], totals["expense"], "", totals["visitors"], totals["bdays"], ""]
    for ci, v in enumerate(summary, 1):
        c = ws2.cell(row=tr, column=ci, value=v)
        c.font, c.fill, c.alignment, c.border = TOT_F, TOT_FL, CENTER, BORDER

    ws2.freeze_panes = "A2"
    buf = io.BytesIO()
    wb2.save(buf)
    buf.seek(0)
    return buf.read()

